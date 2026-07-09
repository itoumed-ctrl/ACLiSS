#!/usr/bin/env python3
"""
ACLiSSマスタ.xlsx（材料シート・オーダ可能項目シート）を、
/admin からアップロードできるCSV形式に変換するスクリプト。

使い方:
  pip install openpyxl
  python3 scripts/convert-master-xlsx.py <材料シートのxlsx> <オーダ可能項目シートのxlsx> <出力先フォルダ>
"""

import csv
import re
import sys
from pathlib import Path

import openpyxl


def convert_containers(xlsx_path: Path, out_path: Path) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    fieldnames = [
        "container_code",
        "vessel",
        "material",
        "dispense_location",
        "dispense_phs",
        "inquiry_dept",
        "inquiry_phs",
        "item_count",
        "collection_amount",
        "representative_item_code",
        "test_summary",
        "has_instruction",
        "instruction_1",
        "instruction_2",
        "instruction_3",
        "notes",
        "image_path_raw",
        "image_source_code",
    ]

    count = 0
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            code = (r[0] or "").strip() if isinstance(r[0], str) else r[0]
            if not code:
                continue
            instruction_1 = r[12] or ""
            instruction_2 = r[13] or ""
            instruction_3 = r[14] or ""
            # 指示有無(列11)は入力ミスの値が混在していることがあるため使わず、
            # 実際に指示文が入っているかどうかで判定する。
            has_instruction = "1" if (instruction_1 or instruction_2 or instruction_3) else "0"

            # 「画像パス」列（列16）の末尾のファイル名が、実際に表示すべき画像を
            # 持つ容器コード。写真が複数の容器コードで使い回されていることがあるため、
            # container_code とは別に記録する（一致しない場合は使い回し）。
            image_path_raw = r[16] or ""
            match = re.search(r"([0-9A-Za-z]+)\.[^.\\/]+$", image_path_raw)
            image_source_code = match.group(1) if match else code
            if image_source_code.upper() == "NI":  # "No Image" のプレースホルダー
                image_source_code = ""

            writer.writerow(
                {
                    "container_code": code,
                    "vessel": r[1] or "",
                    "material": r[2] or "",
                    "dispense_location": r[3] or "",
                    "dispense_phs": r[4] or "",
                    "inquiry_dept": r[5] or "",
                    "inquiry_phs": r[6] or "",
                    "item_count": r[7] if r[7] is not None else "",
                    "collection_amount": r[8] or "",
                    "representative_item_code": r[9] or "",
                    "test_summary": r[10] or "",
                    "has_instruction": has_instruction,
                    "instruction_1": instruction_1,
                    "instruction_2": instruction_2,
                    "instruction_3": instruction_3,
                    "notes": r[15] or "",
                    "image_path_raw": image_path_raw,
                    "image_source_code": image_source_code,
                }
            )
            count += 1
    return count


def convert_test_items(xlsx_path: Path, out_path: Path, valid_container_codes: set[str]) -> tuple[int, list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    fieldnames = ["test_item_code", "test_item_name", "container_code"]
    orphans = []

    count = 0
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            raw_code = r[0]
            code = raw_code.strip() if isinstance(raw_code, str) else raw_code
            if not code:
                continue
            container_code = (r[2] or "").strip() if isinstance(r[2], str) else r[2]
            if container_code and container_code not in valid_container_codes:
                orphans.append(f"{code} ({r[1]}) -> 存在しない容器コード '{container_code}'")
                container_code = ""  # 外部キー制約に違反しないよう空にする

            writer.writerow(
                {
                    "test_item_code": code,
                    "test_item_name": (r[1] or "").strip(),
                    "container_code": container_code or "",
                }
            )
            count += 1
    return count, orphans


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    containers_xlsx = Path(sys.argv[1])
    test_items_xlsx = Path(sys.argv[2])
    out_dir = Path(sys.argv[3])
    out_dir.mkdir(parents=True, exist_ok=True)

    containers_out = out_dir / "containers.csv"
    test_items_out = out_dir / "test_items.csv"

    n_containers = convert_containers(containers_xlsx, containers_out)

    wb = openpyxl.load_workbook(containers_xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    valid_codes = {
        (r[0] or "").strip() if isinstance(r[0], str) else r[0]
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[0]
    }

    n_test_items, orphans = convert_test_items(test_items_xlsx, test_items_out, valid_codes)

    print(f"容器マスタ: {n_containers}件 -> {containers_out}")
    print(f"検査項目マスタ: {n_test_items}件 -> {test_items_out}")
    if orphans:
        print(f"\n注意: 存在しない容器コードを参照している検査項目が{len(orphans)}件ありました（container_codeを空にして出力）:")
        for o in orphans:
            print(" -", o)


if __name__ == "__main__":
    main()
